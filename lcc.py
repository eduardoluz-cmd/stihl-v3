"""
Carregamento do Excel DASHBOARD_LCC.xlsx + cálculo de farol e LCC/ROI.

- carregar_dados_excel(): lê abas Dados + MATRIZ e devolve um DataFrame
  consolidado de 742 ativos.
- farol_v2(): régua de farol conforme proposta do e-mail (Fagner/Diego).
- classificar_tendencia(): rotula slope anual em Queda/Estável/Alta.
- calcular_lcc(): replica as fórmulas de Dados!AC:AV (NPV, CAE, ROI 7 anos).
"""

from __future__ import annotations
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

EXCEL_PATH = Path(__file__).parent / "data" / "DASHBOARD_LCC.xlsx"

HORIZONTE_ANOS = 7
TAXA_PADRAO = 0.10
REDUCAO_PADRAO = 0.80


# =============================================================================
# CARREGAMENTO DO EXCEL
# =============================================================================
def carregar_dados_excel(path: Path | str = EXCEL_PATH) -> pd.DataFrame:
    """Lê Dados + MATRIZ do workbook e devolve DF unificado por TAG.

    Inclui inputs manuais (β, valores, custos) + obsolescência da MATRIZ +
    métricas derivadas (idade, falhas/ano, MTBF nominal).
    """
    path = Path(path)

    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df_dados = pd.read_excel(
            path,
            sheet_name="Dados",
            engine="openpyxl",
            header=0,
        )
        # MATRIZ tem header em duas linhas (4 e 5) com colunas mescladas.
        # Leitura direta via openpyxl: coluna C = TAG, coluna J = Situação.
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["MATRIZ"]
        rows = []
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row,
                                min_col=3, max_col=10, values_only=True):
            tag = row[0]  # col C
            sit = row[7]  # col J
            if tag is None or str(tag).strip() in ("", "nan"):
                continue
            rows.append({"TAG": str(tag).strip(), "Obsolescencia": sit})
        df_matriz = pd.DataFrame(rows)
        wb.close()

    # ---- Limpeza Dados ----
    df = df_dados.rename(columns={
        "TAG": "TAG",
        "Denominação": "Denominacao",
        "Nº SAP": "SAP",
        "Local Instalação": "Loc_instalacao",
        "Centro Custo": "Centro_custo",
        "Ano Aquisição": "Ano_aquisicao",
        "Valor de Aquisição Atual": "Valor_aquisicao",
        "Valor Residual Atual": "Valor_residual",
        "Custo de Manut. Total": "Custo_manut_total",
        "Beta": "Beta",
        "Tempo de Missão (Horas Trab Ano)": "Tempo_missao",
        "Tempo Total": "Tempo_total",
        "Qtd Falhas Totais": "Qtd_falhas",
        "Custo Esperado por Corretiva": "Custo_corretiva",
        "Custo Esperado Preventiva Mensal": "Custo_preventiva_mensal",
        "Taxa": "Taxa",
        "Acréscimo Aquisição": "Inflacao_acrescimo",
        "Red. Custo Manut.": "Reducao_manut",
        "Mini Fab": "Mini_fab",
    })

    keep = [
        "TAG", "Denominacao", "SAP", "Loc_instalacao", "Centro_custo",
        "Ano_aquisicao", "Valor_aquisicao", "Valor_residual",
        "Custo_manut_total", "Beta", "Tempo_missao", "Tempo_total",
        "Qtd_falhas", "Custo_corretiva", "Custo_preventiva_mensal",
        "Taxa", "Inflacao_acrescimo", "Reducao_manut", "Mini_fab",
    ]
    df = df[[c for c in keep if c in df.columns]].copy()

    df["TAG"] = df["TAG"].astype(str).str.strip()
    df = df[df["TAG"].notna() & (df["TAG"] != "") & (df["TAG"] != "nan")]

    # Tudo o que é texto: força string para o Arrow não engasgar
    for c in ["Denominacao", "SAP", "Loc_instalacao", "Centro_custo", "Mini_fab"]:
        if c in df.columns:
            df[c] = df[c].astype(str).replace({"nan": None, "None": None})

    # Mini Fáb com VLOOKUP quebrado vira "(sem)" para evitar "nan" em cards
    if "Mini_fab" in df.columns:
        df["Mini_fab"] = df["Mini_fab"].fillna("(sem)").replace({"": "(sem)"})

    # Numéricos: tudo o que vem de fórmula vira NaN se inválido
    num_cols = [
        "Ano_aquisicao", "Valor_aquisicao", "Valor_residual",
        "Custo_manut_total", "Beta", "Tempo_missao", "Tempo_total",
        "Qtd_falhas", "Custo_corretiva", "Custo_preventiva_mensal",
        "Taxa", "Inflacao_acrescimo", "Reducao_manut",
    ]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    # Defaults
    df["Taxa"] = df["Taxa"].fillna(TAXA_PADRAO)
    df["Reducao_manut"] = df["Reducao_manut"].fillna(REDUCAO_PADRAO)
    df["Inflacao_acrescimo"] = df["Inflacao_acrescimo"].fillna(0.0)
    df["Valor_residual"] = df["Valor_residual"].fillna(0.0)

    # ---- MATRIZ: obsolescência (Situação) ----
    if not df_matriz.empty:
        df_matriz["Obsolescencia"] = df_matriz["Obsolescencia"].astype(str).str.strip()
        df_matriz.loc[
            df_matriz["Obsolescencia"].isin(["", "nan", "None"]),
            "Obsolescencia",
        ] = "Não há"
        df = df.merge(df_matriz, on="TAG", how="left")
    else:
        df["Obsolescencia"] = "Não há"

    df["Obsolescencia"] = df["Obsolescencia"].fillna("Não há")

    # ---- Derivadas ----
    ano_corrente = 2026
    df["Idade_anos"] = (ano_corrente - df["Ano_aquisicao"]).clip(lower=0)

    # Falhas por ano = qtd_total × (tempo_missao / tempo_total)
    mask_ok = (df["Tempo_total"] > 0) & (df["Tempo_missao"] > 0) & (df["Qtd_falhas"] > 0)
    df["Falhas_por_ano"] = np.where(
        mask_ok,
        df["Qtd_falhas"] * df["Tempo_missao"] / df["Tempo_total"],
        np.nan,
    )

    # MTBF nominal (h)
    df["MTBF_h"] = np.where(
        df["Qtd_falhas"] > 0,
        df["Tempo_total"] / df["Qtd_falhas"],
        np.nan,
    )

    # Custo médio anual de manutenção (até hoje)
    df["Custo_anual_medio"] = np.where(
        df["Idade_anos"] > 0,
        df["Custo_manut_total"] / df["Idade_anos"],
        df["Custo_manut_total"],
    )

    return df.reset_index(drop=True)


# =============================================================================
# FAROL V2 — proposta do e-mail
# =============================================================================
FAROL_NIVEIS = ["verde", "amarelo", "vermelho"]
FAROL_EMOJI = {"verde": "🟢", "amarelo": "🟡", "vermelho": "🔴"}
FAROL_LABEL = {"verde": "Manter", "amarelo": "Acompanhar", "vermelho": "Substituir"}


def classificar_beta_v2(beta: float) -> tuple[str, str]:
    """Retorna (rótulo da fase, cor base do farol) — régua do e-mail."""
    if pd.isna(beta):
        return "Sem dado", "amarelo"
    if beta < 0.9:
        return "Falhas prematuras / instabilidade", "amarelo"
    if beta <= 1.2:
        return "Aleatório controlado", "verde"
    if beta <= 1.5:
        return "Início de desgaste", "amarelo"
    return "Desgaste claro", "vermelho"


def _nivel(cor: str) -> int:
    return FAROL_NIVEIS.index(cor)


def _cor(nivel: int) -> str:
    return FAROL_NIVEIS[min(max(nivel, 0), 2)]


def farol_v2(
    beta: float,
    obsolescencia: str = "Não há",
    tend_mtbf: str = "N/D",
    tend_custo: str = "N/D",
) -> dict:
    """Régua do e-mail (Fagner/Diego):

    - Base por β (4 faixas)
    - Obsolescência Total → força vermelho
    - Obsolescência Parcial → agrava +1
    - MTBF "Queda" → agrava +1
    - Custo "Crescente" → agrava +1
    - Nunca melhora por tendência positiva; cap em vermelho.
    """
    fase, cor_base = classificar_beta_v2(beta)
    nivel = _nivel(cor_base)
    agravantes = []

    obs = (obsolescencia or "").strip()

    if obs.lower().startswith("total"):
        return {
            "cor": "vermelho",
            "fase_beta": fase,
            "cor_base": cor_base,
            "agravantes": ["Obsolescência Total (override)"],
        }

    if obs.lower().startswith("parcial"):
        nivel += 1
        agravantes.append("Obsolescência Parcial")

    if tend_mtbf == "Queda":
        nivel += 1
        agravantes.append("MTBF em queda")

    if tend_custo == "Crescente":
        nivel += 1
        agravantes.append("Custo crescente")

    return {
        "cor": _cor(nivel),
        "fase_beta": fase,
        "cor_base": cor_base,
        "agravantes": agravantes,
    }


# =============================================================================
# TENDÊNCIAS (rótulos Queda/Estável/Alta e Crescente/Estável/Decrescente)
# =============================================================================
def classificar_tendencia(
    serie_anual: pd.Series,
    tipo: str = "mtbf",
    limiar_rel: float = 0.05,
) -> str:
    """Classifica o slope anual relativo à média.

    tipo="mtbf"  → Queda / Estável / Alta
    tipo="custo" → Crescente / Estável / Decrescente
    """
    s = serie_anual.dropna()
    if len(s) < 3:
        return "N/D"

    x = np.arange(len(s), dtype=float)
    y = s.values.astype(float)

    media = float(np.mean(y))
    if media == 0:
        return "Estável"

    slope = float(np.polyfit(x, y, 1)[0])
    slope_rel = slope / abs(media)

    if tipo == "mtbf":
        if slope_rel < -limiar_rel:
            return "Queda"
        if slope_rel > limiar_rel:
            return "Alta"
        return "Estável"
    else:  # custo
        if slope_rel > limiar_rel:
            return "Crescente"
        if slope_rel < -limiar_rel:
            return "Decrescente"
        return "Estável"


# =============================================================================
# CÁLCULO DE LCC / ROI — replicando Excel (Dados!AC:AV + Bruta)
# =============================================================================
def calcular_lcc(
    valor_aquisicao: float,
    valor_residual: float,
    custo_manut_total: float,
    falhas_por_ano: float,
    custo_corretiva: float,
    custo_preventiva_mensal: float,
    taxa: float = TAXA_PADRAO,
    reducao_manut: float = REDUCAO_PADRAO,
    inflacao_acrescimo: float = 0.0,
    horizonte: int = HORIZONTE_ANOS,
) -> dict:
    """Replica Dados!AC:AV + Bruta (LCC, NPV, CAE, ROI, Payback).

    Convenção do Excel: custos são negativos (saídas de caixa); investimento
    novo entra como CAPEX negativo no ano 0. Falhas crescem linearmente —
    "P, 2P, 3P..." — então o incremento por ano é constante (= P).

    custo_ano_k = -[(P × W) + (X × 12)] × (1 + taxa)^k
    custo_novo_k = custo_ano_k × (1 - reducao)
    """
    anos = np.arange(1, horizonte + 1)
    P = float(falhas_por_ano or 0.0)
    W = float(custo_corretiva or 0.0)
    X = float(custo_preventiva_mensal or 0.0)
    Y = float(taxa or TAXA_PADRAO)
    AA = float(reducao_manut if reducao_manut is not None else REDUCAO_PADRAO)
    Z = float(inflacao_acrescimo or 0.0)
    I = float(valor_aquisicao or 0.0)
    J = float(valor_residual or 0.0)

    # Cenário ATUAL — manter
    custo_anual_base = (P * W) + (X * 12)
    fluxo_atual = -custo_anual_base * (1 + Y) ** anos  # negativo = saída

    # Cenário NOVO — substituir
    fluxo_novo = fluxo_atual * (1 - AA)
    capex_novo = -I * (1 + Z)

    # Valor presente líquido (NPV do Excel: desconta o 1º fluxo no t=1)
    fatores = (1 + Y) ** anos
    npv_atual = float(np.sum(fluxo_atual / fatores))
    npv_novo = float(np.sum(fluxo_novo / fatores))

    # Potencial de Ganho = (NPV novo + CAPEX − residual recuperado) − NPV atual
    potencial_ganho = (npv_novo + capex_novo - J) - npv_atual

    # ROI estilo Excel (Dados.AU/AV): em mil R$
    profit = (
        (np.mean(fluxo_atual) / 1000 - np.mean(fluxo_novo) / 1000)
        + ((capex_novo / 1000) * -1) / horizonte
        + (J / 1000) / horizonte
    )
    investimento = (J / 1000) + ((-capex_novo) / 1000)
    roi = profit / investimento if investimento > 0 else np.nan

    # CAE — custo anual equivalente (Bruta!C15)
    def _cae(npv, n):
        if Y <= 0:
            return npv / n
        f = ((1 + Y) ** n) * Y / (((1 + Y) ** n) - 1)
        return npv * f

    cae_atual = _cae(npv_atual, horizonte)
    cae_novo = _cae(npv_novo + capex_novo - J, horizonte)

    # Payback simples (anos para investimento ser pago pela economia anual)
    economia_anual = custo_anual_base * AA  # economia média (sem desconto)
    investimento_liq = (-capex_novo) - J
    if economia_anual > 0:
        payback = investimento_liq / economia_anual
    else:
        payback = np.nan

    # Soma nominal dos custos (= Dados.AS / AT)
    custo_esperado_atual = float(np.sum(fluxo_atual))
    custo_esperado_novo = float(np.sum(fluxo_novo))

    return {
        "fluxo_atual": fluxo_atual,
        "fluxo_novo": fluxo_novo,
        "capex_novo": capex_novo,
        "valor_residual": J,
        "npv_atual": npv_atual,
        "npv_novo": npv_novo,
        "custo_esperado_atual": custo_esperado_atual,
        "custo_esperado_novo": custo_esperado_novo,
        "potencial_ganho": potencial_ganho,
        "profit": profit,
        "roi": roi,
        "cae_atual": cae_atual,
        "cae_novo": cae_novo,
        "payback_anos": payback,
        "anos": anos,
    }
